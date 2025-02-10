"""
#读写csv文件
import csv
import random
with open('score.csv','w',newline='',encoding='utf-8') as file:
    my_writer = csv.writer(file)
    # my_writer=csv.writer(file,delimiter='|',quoting=csv.QUOTE_ALL)
    my_writer.writerow(['姓名','语文','数学','英语','科学'])
    names=['小李','小张','小陈']
    for name in names:
        scores=[random.randint(40,100) for _ in range(4)]
        scores.insert(0,name)
        my_writer.writerow(scores)

import csv
with open('score.csv','r',encoding='utf-8') as file:
    my_reader=csv.reader(file,delimiter='|')
    for data_list in my_reader:
        print(my_reader.line_num,end='\t')
        for elem in data_list:
            print(elem,end='\t')
        print()
"""
from asyncore import write
from pydoc import browse
from types import NoneType

"""
#读写Excel文件
import openpyxl
import datetime
work_book=openpyxl.load_workbook('2022年股票数据.xlsx')
sheet_names=work_book.sheetnames
print(sheet_names)
sheet=work_book.worksheets[0]
print(sheet.dimensions)
print(sheet.max_row,sheet.max_column)
print(sheet.cell(3,3).value)
print(sheet['A3'].value)
print(sheet['A2:C5'])
for row in sheet['A2:C5']:
    values=[cell.value for cell in row]
    print(values)
for row in range(2,sheet.max_row+1):
    for col in 'ABCDEF':
        value=sheet[f'{col}{row}'].value
        if type(value)==datetime.datetime:
            print(value.strftime('%Y年%m月%d日'),end='\t')
        elif type(value)==int:
            print(f'{value:<8d}',end=' ')
        elif type(value)==float:
            print(f'{value:.4f}',end=' ')
        else:
            print(value)
    print()

import openpyxl
import random
wb=openpyxl.Workbook()
sheet=wb.active
sheet.title='期末成绩'
titles=('姓名','语文','数学')
for col_index,title in enumerate(titles):
    sheet.cell(1,col_index+1,title)
names=['小李','小张','小胡']
for row_index,name in enumerate(names):
    sheet.cell(row_index+2,1,name)
    for col_index in range(2):
        sheet.cell(row_index+2,col_index+2,random.randint(40,100))
wb.save('期末成绩.xlsx')
"""
"""
#调整表格样式
import openpyxl
from openpyxl.styles import Alignment,Font,Border,Side
work_book=openpyxl.load_workbook('期末成绩.xlsx')
sheet=work_book.worksheets[0]
alignment=Alignment(horizontal='center',vertical='center')
side=Side(color='ff7f50',style='mediumDashed')
sheet.row_dimensions[1].height=30
sheet.column_dimensions['D'].width=15
sheet['D1']='平均分'
sheet.cell(1,4).font=Font(size=18,bold=True,color='ff1493',name='华文楷体')
sheet.cell(1,4).alignment=alignment
sheet.cell(1,4).border=Border(left=side,right=side,top=side,bottom=side)
for row_index in range(2,5):
    Sum=0
    for col_index in range(2,4):
        Sum=Sum+sheet.cell(row_index,col_index).value
    sheet.cell(row_index,4).value=Sum/2.0
    sheet.cell(row_index,4).font=Font(size=18,bold=True,color='4169e1',italic=True)
    sheet.cell(row_index,4).border=Border(left=side,right=side,top=side,bottom=side)
    sheet.cell(row_index,4).alignment=alignment
work_book.save('期末成绩.xlsx')
"""

#生成统计图表
import openpyxl
from openpyxl.chart import BarChart,Reference
wb=openpyxl.Workbook(write_only=True)
sheet=wb.create_sheet()
rows = [
    ('类别', '销售A组', '销售B组'),
    ('手机', 40, 30),
    ('平板', 50, 60),
    ('笔记本', 80, 70),
    ('外围设备', 20, 10),
]
for row in rows:
    sheet.append(row)
chart=BarChart()
chart.type='col'
chart.style=8
chart.title='销量统计图'
chart.x_axis.title='类别'
chart.y_axis.title='销量'
data=Reference(sheet,min_col=2,max_col=3,min_row=1,max_row=5)
cats=Reference(sheet,min_col=1,min_row=2,max_row=5)
chart.add_data(data,titles_from_data=True)
chart.set_categories(cats)
chart.shape=4
sheet.add_chart(chart,'A10')
wb.save('销量统计.xlsx')

